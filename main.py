import requests
import json
import os
from openpyxl import Workbook
import datetime

class ClickUpExporter:
    """
    A class for exporting tasks from multiple ClickUp lists into a single Excel workbook.
    """

    def __init__(self, api_token, team_id, list_ids, output_path, output_filename_format):
        """
        Initialize the ClickUpExporter object.

        Args:
            api_token (str): The ClickUp API token.
            team_id (str): The ClickUp team ID.
            list_ids (dict): A dictionary with sheet names as keys and ClickUp list IDs as values.
            output_path (str): Directory path to save the Excel file.
            output_filename_format (str): Filename format for output, e.g., 'backup_{date}.xlsx'.
        """
        self.api_token = api_token
        self.team_id = team_id
        self.list_ids = list_ids
        self.output_path = output_path
        self.output_filename_format = output_filename_format

    def _get_tasks_from_list(self, list_id):
        """
        Retrieve tasks from a specific ClickUp list.

        Args:
            list_id (str): The ClickUp list ID.

        Returns:
            list: A list of task dictionaries.
        """
        url = f"https://api.clickup.com/api/v2/list/{list_id}/task"
        headers = {"Authorization": self.api_token}
        params = {"include_subtasks": "true"}
        try:
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            return response.json().get("tasks", [])
        except requests.exceptions.RequestException as e:
            print(f"Error retrieving tasks from list with ID {list_id}: {e}")
            return []

    def _get_field_value(self, field):
        """
        Retrieves and formats the value of a custom field in a ClickUp task.

        Args:
            field (dict): A dictionary representing the custom field.

        Returns:
            str: The formatted value of the field as a string. Handles various ClickUp field types.
        """
        field_name = field.get("name")
        field_type = field.get("type")
        field_value = field.get("value")

        if field_value is None:
            return ""

        if field_type in ["text", "short_text", "email", "phone", "url"]:
            return str(field_value)

        elif field_type in ["number", "rating", "auto_increment"]:
            return str(field_value)

        elif field_type == "checkbox":
            return "Yes" if field_value else "No"

        elif field_type == "dropdown":
            # New Correct Logic:
            if isinstance(field_value, str):
                options = field.get("type_config", {}).get("options", [])
                for option in options:
                    if option.get("id") == field_value:
                        return option.get("name", "")
                # fallback - if not found
                return field_value
            elif isinstance(field_value, dict):
                return field_value.get("name", "")
            else:
                return str(field_value)

        elif field_type == "labels" or field_type == "multi_select":
            if isinstance(field_value, list):
                names = []
                options = field.get("type_config", {}).get("options", [])
                for item in field_value:
                    if isinstance(item, dict) and "name" in item:
                        names.append(item["name"])
                    else:
                        # search by ID
                        for option in options:
                            if option.get("id") == item:
                                names.append(option.get("name", ""))
                return ", ".join(names)
            return ""

        elif field_type == "date" or field_type == "time":
            return str(field_value / 1000)  # ms to seconds

        elif field_type == "users":
            if isinstance(field_value, list):
                return ", ".join(user.get("username", "") for user in field_value if isinstance(user, dict))
            else:
                return ""

        elif field_type == "location":
            if isinstance(field_value, dict):
                return field_value.get("name", "")
            return ""

        elif field_type == "relationship":
            if isinstance(field_value, list):
                return ", ".join(linked_task.get("name", "") for linked_task in field_value if linked_task)
            return ""

        elif field_type == "formula":
            if isinstance(field_value, dict):
                return field_value.get("text", "")
            return ""

        elif field_type in ["created_by", "updated_by"]:
            if isinstance(field_value, dict):
                return field_value.get("username", "")
            return ""

        else:
            return str(field_value)

    def _process_task(self, task, custom_field_names):
        """
        Processes a single task, extracting its name and all custom fields.

        Args:
            task (dict): The ClickUp task dictionary.
            custom_field_names (dict): Dictionary of all unique custom fields names across tasks.

        Returns:
            list: A list with the task name and custom field values.
        """
        row_data = [task.get("name", "")]
        custom_field_values = {name: "" for name in custom_field_names.keys()}

        for field in task.get("custom_fields", []):
            field_name = field.get("name")
            if field_name in custom_field_values:
                custom_field_values[field_name] = self._get_field_value(field)

        row_data.extend(custom_field_values.values())
        return row_data

    def export_to_excel(self):
        """
        Exports the tasks from specified lists into an Excel file.
        Each ClickUp list is saved into a separate Excel sheet.
        """
        try:
            workbook = Workbook()

            for sheet_name, list_id in self.list_ids.items():
                tasks = self._get_tasks_from_list(list_id)
                if not tasks:
                    print(f"No tasks found for export from list '{sheet_name}'.")
                    continue

                sheet = workbook.create_sheet(title=sheet_name)
                headers = ["Task Name"]
                custom_field_names = {}

                for task in tasks:
                    for field in task.get("custom_fields", []):
                        custom_field_names[field.get("name")] = True

                headers.extend(custom_field_names.keys())
                sheet.append(headers)

                for task in tasks:
                    row_data = self._process_task(task, custom_field_names)
                    sheet.append(row_data)

            # Remove the default 'Sheet' if more sheets have been created
            if len(workbook.sheetnames) > 1 and "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]

            output_filename = self.output_filename_format.format(date=datetime.date.today().strftime('%Y-%m-%d'))
            filepath = os.path.join(self.output_path, output_filename)
            workbook.save(filepath)
            print(f"Data from lists saved to: {filepath}")

        except Exception as e:
            print(f"Error during Excel file saving: {e}")
        finally:
            try:
                workbook.close()
            except Exception as e:
                print(f"Error closing the Excel file: {e}")

if __name__ == "__main__":
    # Configuration
    CLICKUP_API_TOKEN = ""   # Replace with your ClickUp API token
    CLICKUP_TEAM_ID = ""   # Replace with your ClickUp team ID

    LIST_IDS = {
        "Biuro": "",
        "MP Energy": "",
        "Wykonawstwo": "",
        "Kontakty OSD": "",
    }

    # Reading the path from BackupPath.txt
    with open('BackupPath.txt', 'r') as file:
        content = file.read()
        print(f"Saving backup to: {content}")

    OUTPUT_PATH = content
    OUTPUT_FILENAME = f"0000 Kontakty_BACKUP[{{date}}].xlsx"

    exporter = ClickUpExporter(CLICKUP_API_TOKEN, CLICKUP_TEAM_ID, LIST_IDS, OUTPUT_PATH, OUTPUT_FILENAME)
    exporter.export_to_excel()
